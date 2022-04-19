from sqlite3 import Row
from openpyxl import Workbook,load_workbook

class courseSort:
    def __init__(self):
        filename = "Udemy Free (en) - (popularity).xlsx"
        wb = load_workbook(filename)
        ws = wb.active

        rows_number = ws.max_row
        columns_number = ws.max_column

        # self.titleSort(ws, rows_number)
        # self.authorSort(ws, rows_number)
        # self.shortInfo(ws, rows_number)

        keyword = "Course"
        self.searchKeyword(ws, rows_number, keyword.lower())

    def titleSort(self, ws, rows_number):
        print("-"*100)
        print(" Title "*10)
        print("-"*100)
        for row in range(2, rows_number):   # Kurs İsimleri Listesi      
            print(row-1 , " | " + str(ws.cell(row,2).value))
        print("-"*100)
    
    def authorSort(self, ws, rows_number):
        print("-"*100)
        print(" Author "*10)
        print("-"*100)
        for row in range(2, rows_number):   # Yazarlar Listesi      
            print(row-1 , " | " + str(ws.cell(row,4).value))
        print("-"*100)

    def shortInfo(self, ws, rows_number):
        print("-"*100)
        print(" Short "*10)
        print("-"*100)

        for i in range(2, rows_number):
            print(" | " + str(ws[f'B{i}'].value) + " | ",end="")
            print(" | " + str(ws[f'D{i}'].value) + " | ",end="")
            print(" | " + str(ws[f'F{i}'].value) + " | ",end="")
            print(" | " + str(ws[f'I{i}'].value) + " | ",end="")
            print()

    def searchKeyword(self, ws, rows_number, keyword):
        print("-"*100)
        print(" Short "*10)
        print("-"*100)
        temp_keyword = ""

        for i in range(2, 15):
            for satir in ws[f'B{i}':f'C{i}']:
                for hucre in satir:
                    text = hucre.value.lower()

                    if keyword in text:
                        chech = True
                        print("kelime bulundu")
                        print(" | " + str(ws[f'B{i}'].value) + " | ",end="")
                        print(" | " + str(ws[f'C{i}'].value) + " | ",end="")
                        print()
        
        if(chech==True):
            print("Basarili")
        else:
            print("Aradaginiz kelime bulunamadi.")


course = courseSort()
