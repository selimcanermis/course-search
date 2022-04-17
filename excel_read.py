from openpyxl import Workbook,load_workbook #pip install openpyxl

wb = load_workbook("Udemy Free (en) - (popularity).xlsx")
ws = wb.active

print(wb.sheetnames)
print(ws)

ws = wb["Sheet1"]       #Sheet1 aktif edildi artık orada çalışır.
print(ws)

# 1. Kullanım
print(ws["D2"].value)        # Chris DeLeon

# 2. Kullanım
print(ws.cell(2,4).value)    # Chris DeLeon

for satir in range(2,21):   # Yazarlar Listesi      
    print(" | " + str(ws.cell(satir,4).value))

print("*"*100)

for satir in range(2,21):
    for sutun in range(2,5):        
        print(" | " + str(ws.cell(satir,sutun).value) + " | ",end="")
    print()

print("*"*100)

for satir in ws['A2':'C4']:
    for hucre in satir:
        print(" | " + str(hucre.value) + " | ",end="")
    print()

print("*"*100)

print("Rows: ", ws.max_row)
print("Columns: ", ws.max_column)

print("*"*100)

"""
for satir in range(1,ws.max_row+1):
    for sutun in range(1,ws.max_column+1):
        print(" | " + str(ws.cell(satir,sutun).value) + " | ",end="")
    print()

print("*"*100)
"""

"""
# * YENİ EXCEL DOSYASI AÇMA

wb = Workbook()

ws = wb.active
ws.title = "İlk Çalışma Alanı"
ws = wb.create_sheet("Posta Kodları")
ws = wb.create_sheet("Ülkeler")

print(wb.sheetnames)     # ['İlk Çalışma Alanı', 'Posta Kodları', 'Ülkeler']

wb.save("dosyaAdi.xlsx")

ws['A1'] = 42
ws['B3'] = "Merhaba"
ws.append([1, 2, 3, "Merhaba", "Dünya"])  # Sıradaki satıra sırasıyla dizi elemanlarını ekler

ws.sheet_properties.tabColor = "1072BA"

ws.insert_rows(satirno)
ws.delete_rows(satirno)
ws.insert_cols(sutunno)
ws.delete_cols(sutunno)

ws["P2"] = "=AVERAGE(H2:H100)"

kalin = Font(bold=True)
buyuk_kirmizi_yazi = Font(color=colors.RED, size=20)
hizala_orta = Alignment(horizontal="center")
cift_kenarlik = Side(border_style="double")
kenarlik = Border(top=cift_kenarlik, right=cift_kenarlik, bottom=cift_kenarlik, left=cift_kenarlik)

ws["A2"].font = kalin
ws["A3"].font = buyuk_kirmizi_yazi
ws["A4"].alignment = hizala_orta
ws["A5"].border = kenarlik

"""

"""
#! RESİM EKLEME
from openpyxl.drawing.image import Image

logo = Image("logo.png")
logo.height = 150
logo.width = 150

ws.add_image(logo, "A3")
"""